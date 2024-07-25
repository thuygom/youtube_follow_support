import openai
from openpyxl import load_workbook, Workbook
from sklearn.metrics import accuracy_score, f1_score
from collections import Counter
import time

# OpenAI API 키 설정
openai.api_key = 'key'
context = {
    'oakQvwCbvr8':'스캠 코인 사기에 대한 사과영상', '-HZeqsIgGHo':'스캠 코인 사기에 대한 사과영상', 'nAy-7zuCVQs':'스캠 코인 사기에 대한 사과영상', 'QojVuirFx58':'스캠 코인 사기에 대한 사과영상', 'ibI5OOZXSj8':'남자들끼리의 대화방식에 대한 코미디 영상',
    'g5KDoSqT24Q':'여자들이 싫어하는 여름신발에 대한 소개영상', 'mnn1_yu0aDQ':'소장품 3000벌을 소개해주고 구독자 중 선물해주는 영상', 'pp_C0MGj9ZM':'유니클로에서 세일할때 사면 좋은 필수템에 대한 소개', '_Otk-iMD_X0':'남자들의 여름 옷 코디 꿀팁', '_HZ63R-8z4E':'실버 악세서리 소개',
    'yLlKOd3I8CA':'한국 축구 국대 감독 홍명보에 대한 놀란 반응영상', 'BYWO-z-4tfo':'친구 3명과 게임을 하는 영상', 'uNq7RMRwIHs':'친구 3명과 같이 게임하는 영상', 'ZLXz98YW_U0':'게임 캐릭터가 상향을 받아 기쁜 영상', 'qkXc1M3d7g4':'친구와 장난치며 노는 영',
    '-rHqqCxjhM4':'03년생이 좋아하는 플레이리스트', 'FGjrtBFgTXY':'하이키 노래 소개 영상', 'TOSrdHy5KRc':'밴드 노래 소개영상', 'wdUu3XQK8cE':'DPR IAN의 노래에 대한 플레이리스트', 'LamRCcz4zqg':'eldon-fraud 노래에 대한 소개영',
    'ahcPfSLbT-M':'안드레진이 일으킨 사회적 물의에 대한 비판영상', '8l4GZ4datyM':'달씨 유튜버의 전세사기 시도에 대한 비판영상', '7I790Er-zkc':'허웅 전여친 논란에 대한 반전내용에 대한 영상', '8SJs1Cg7hpU':'허웅 전여친 논란에 비판영상', 'VWmWScovllY':'달씨 유튜버의 전세사기 해명에 대한 비판영'
    }
youtuber = {
    'oakQvwCbvr8':'오킹', '-HZeqsIgGHo':'오킹', 'nAy-7zuCVQs':'오킹', 'QojVuirFx58':'오킹', 'ibI5OOZXSj8':'오킹',
    'g5KDoSqT24Q':'깡 스타일리스트', 'mnn1_yu0aDQ':'깡 스타일리스트', 'pp_C0MGj9ZM':'깡 스타일리스트', '_Otk-iMD_X0':'깡 스타일리스트', '_HZ63R-8z4E':'깡 스타일리스트',
    'yLlKOd3I8CA':'동수칸', 'BYWO-z-4tfo':'동수칸', 'uNq7RMRwIHs':'동수칸', 'ZLXz98YW_U0':'동수칸', 'qkXc1M3d7g4':'동수칸',
    '-rHqqCxjhM4':'때잉', 'FGjrtBFgTXY':'때잉', 'TOSrdHy5KRc':'때잉', 'wdUu3XQK8cE':'때잉', 'LamRCcz4zqg':'때잉',
    'ahcPfSLbT-M':'뻑가', '8l4GZ4datyM':'뻑가', '7I790Er-zkc':'뻑가', '8SJs1Cg7hpU':'뻑가', 'VWmWScovllY':'뻑가'
    }

# 감정 문자열을 OpenAI 엔진에 맞게 매핑
label_map = {'행복': 'happy', '혐오': 'disgust', '공포': 'fear', '슬픔': 'sadness', '놀람': 'surprise', '분노': 'anger', '중립': 'neutral'}
reverse_label_map = {v: k for k, v in label_map.items()}

def load_dataset(file_path):
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active  # 현재 활성화된 시트를 가져옴
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        exit(1)
    
    # 'sentence' 열만 가져옴
    data_list = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            sentence = str(row[0])
            data_list.append([sentence])
    
    return data_list  # 전체 데이터 반환

def load_video_ids(file_path):
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active  # 현재 활성화된 시트를 가져옴
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        exit(1)
    
    # 'video_id' 열만 가져옴 (5번째 열이므로 인덱스 4)
    video_ids = []
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        if row[4] is not None:
            video_id = str(row[4])
            video_ids.append(video_id)
    
    return video_ids

# OpenAI API를 사용하여 감정 분석을 수행하는 함수 정의
def predict_sentiment(context, text):
    try:
        messages = [
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": f"'{context}' 문맥을 기반으로 '{text}'를 행복, 웃김, 슬픔, 분노, 중립 5가지 감정중 하나로 라벨링 후 감정 라벨을 반환하라. "}
        ]
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=messages,
            max_tokens=50,
            stop=["\n", "감정 분석 결과:"],
        )
        # OpenAI API의 응답에서 감정 라벨 추출
        predicted_label = response.choices[0].message['content'].strip().lower()
        return predicted_label
    except openai.error.APIError as e:
        print(f"OpenAI API 호출 오류: {e}")
        return "API 호출 오류"

# OpenAI API를 사용하여 감정 분석을 수행하는 함수 정의
def predict_object(context, youtuber, text, num_predictions=5):
    try:
        messages = [
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": f"문맥: {context}. 유튜버: {youtuber}."},
            {"role": "user", "content": "다음 문장이 얘기하는 대상이 context, youtuber, other중 어떤 것인지 선택하세요. 예: '안드레진 미친놈이네' -> context"},
            {"role": "user", "content": f"문장: '{text}'"}
        ]
        results = []
        for _ in range(num_predictions):
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=messages,
                max_tokens=50,
            )
            predicted_label = response.choices[0].message['content'].strip().lower()
            results.append(predicted_label)
        most_common_label = Counter(results).most_common(1)[0][0]
        return most_common_label
    except openai.error.APIError as e:
        print(f"OpenAI API 호출 오류: {e}")
        return "API 호출 오류"
    
# 예측만 수행하고 싶은 경우:
valid_data_file_path = '../xlsx/crawling_auto0709.xlsx'  # 파일 경로에 맞게 수정해주세요
wb = load_workbook(filename=valid_data_file_path)
ws = wb.active

def evaluate_model(file_path, start, end):
    data_list = load_dataset(file_path)
    video_ids = load_video_ids(file_path)
    sentences = [item[0] for item in data_list]

    if len(sentences) != len(video_ids):
        print("문장 수와 비디오 ID 수가 일치하지 않습니다.")
        return

    # 'emotion'와 'object' 열 추가
    if 'emotion' not in ws.columns:
        ws.append(['sentence', 'emotion', 'object'])

    for i, (sentence, video_id) in enumerate(zip(sentences[start:end], video_ids[start:end]), start=start + 2):
        context_info = context.get(video_id, "기타")
        youtuber_info = youtuber.get(video_id, "기타")

        # 감정 예측
        predicted_sentiment = predict_sentiment(sentence, context_info)
        ws.cell(row=i, column=6, value=predicted_sentiment)  # 감정 결과는 6번째 열에 기록

        # 대상 예측
        predicted_object = predict_object(sentence, context_info, youtuber_info)
        ws.cell(row=i, column=7, value=predicted_object)  # 대상 결과는 7번째 열에 기록
        print(f"원문: {sentence}, 예측 감정: {predicted_sentiment}, 대상: {predicted_object}")  # 예측 결과 출력
    time.sleep(2)  # 2초 동안 대기
    wb.save('../xlsx/labeling_0709.xlsx')

# 실행 횟수
num_executions = 25

# 총 데이터 개수
total_data_count = len(load_dataset(valid_data_file_path))

# 한 번에 처리할 데이터 개수
chunk_size = 100

for execution_index in range(num_executions):
    
    # 현재 실행할 데이터의 시작 인덱스
    start_index = execution_index * chunk_size

    # 현재 실행할 데이터의 끝 인덱스
    end_index = min((execution_index + 1) * chunk_size, total_data_count)
    
    print(f"실행 횟수: {execution_index + 1} {start_index} {end_index} {total_data_count}")

    # evaluate_model 함수 실행
    evaluate_model(valid_data_file_path, start_index, end_index)
