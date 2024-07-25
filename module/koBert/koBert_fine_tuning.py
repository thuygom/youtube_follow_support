import torch
from transformers import BertTokenizer, BertForSequenceClassification, Trainer, TrainingArguments
from torch.utils.data import Dataset
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, f1_score

# 학습 데이터 파일 경로
train_file_path = './sampled_train_data.xlsx'

# KoBERT 모델 및 토크나이저 로딩
model_name = "monologg/kobert"
tokenizer = BertTokenizer.from_pretrained(model_name)
num_labels = 7  # 예시로 7개의 감정 클래스
model = BertForSequenceClassification.from_pretrained(model_name, num_labels=num_labels)

# GPU가 있을 경우 활용
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)

# 데이터셋 클래스 정의
class CommentDataset(Dataset):
    def __init__(self, data_list, tokenizer, max_length):
        self.data_list = data_list
        self.tokenizer = tokenizer
        self.max_length = max_length
        
    def __len__(self):
        return len(self.data_list)
    
    def __getitem__(self, idx):
        comment, label = self.data_list[idx]
        
        encoding = self.tokenizer(comment, truncation=True, padding='max_length', max_length=self.max_length, return_tensors='pt')
        inputs = {
            'input_ids': encoding['input_ids'].flatten(),
            'attention_mask': encoding['attention_mask'].flatten(),
            'labels': torch.tensor(label, dtype=torch.long)
        }
        return inputs

def load_dataset(file_path):
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active  # 현재 활성화된 시트를 가져옴
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        exit(1)
    
    # 'sentence'와 'emotion' 열을 가져옴
    data_list = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            sentence = str(row[0])
            emotion = str(row[1])
            # 감정 문자열을 숫자로 매핑
            label_map = {'행복': 0, '혐오': 1, '공포': 2, '슬픔': 3, '놀람': 4, '분노': 5, '중립': 6}
            label = label_map[emotion]
            data_list.append([sentence, label])
    
    return data_list

# 학습 데이터셋 로딩
data_list = load_dataset(train_file_path)
print(f"전체 데이터셋 크기: {len(data_list)}")

# 데이터셋을 학습용과 검증용으로 분리 (60% 학습, 40% 검증)
train_data_list, eval_data_list = train_test_split(data_list, train_size=0.6, test_size=0.4, random_state=42)

train_dataset = CommentDataset(train_data_list, tokenizer, max_length=128)
eval_dataset = CommentDataset(eval_data_list, tokenizer, max_length=128)

training_args = TrainingArguments(
    per_device_train_batch_size=128,
    per_device_eval_batch_size=32,
    num_train_epochs=5,
    logging_dir='./logs',
    logging_steps=100,
    evaluation_strategy="epoch",  # 에포크마다 검증
    save_steps=500,  # 500 steps 마다 모델 저장
    output_dir='./results2',  # 모델 저장 디렉토리
    overwrite_output_dir=True,
)


# 정확도와 F1 score를 계산하는 함수 정의
def compute_metrics(pred):
    labels = pred.label_ids
    preds = pred.predictions.argmax(-1)
    acc = accuracy_score(labels, preds)
    f1 = f1_score(labels, preds, average='weighted')  # weighted average F1 score
    return {
        'accuracy': acc,
        'f1_score': f1,
    }


# Trainer 초기화 및 Fine-tuning
trainer = Trainer(
    model=model,
    args=training_args,
    train_dataset=train_dataset,
    eval_dataset=eval_dataset,
    compute_metrics=compute_metrics,
)

# 모델 가중치 일부 확인
print("모델 가중치 확인:")
print(model.bert.encoder.layer[0].attention.self.query.weight[:5])  # 일부 가중치 출력

# Fine-tuning 진행
trainer.train()

# 학습된 모델 및 토크나이저 저장
trainer.save_model('./results2')
tokenizer.save_pretrained('./results2')
print("모델과 토크나이저가 ./results2 디렉토리에 저장되었습니다.")

# Fine-tuning된 모델을 사용하여 감정 분석 수행 함수 정의
def predict_sentiment(model, tokenizer, text, max_length=128):
    encoded_input = tokenizer(text, truncation=True, padding='max_length', max_length=max_length, return_tensors='pt')
    input_ids = encoded_input['input_ids'].to(device)
    attention_mask = encoded_input['attention_mask'].to(device)
    with torch.no_grad():
        output = model(input_ids=input_ids, attention_mask=attention_mask)
    predicted_label = torch.argmax(output.logits[0]).item()
    return predicted_label

# 감정 라벨의 역매핑 딕셔너리
reverse_label_map = {0: '행복', 1: '혐오', 2: '공포', 3: '슬픔', 4: '놀람', 5: '분노', 6: '중립'}

# 결과 파일에 감정 라벨 추가
def add_labels_to_excel(file_path, tokenizer, model):
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        exit(1)

    # 데이터가 있는 열을 확인하고 데이터 읽기
    column_idx = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'Processed_Comments':  # 또는 다른 식별자를 사용할 수 있음
            column_idx = col
            break
    
    if column_idx is None:
        print("Cannot find 'Processed_Comments' column in the Excel file.")
        exit(1)
    
    sentences = []
    for row in ws.iter_rows(min_row=2, max_col=column_idx, values_only=True):
        if row[0] is not None:
            sentences.append(str(row[0]))
        else:
            sentences.append("")  # 빈 셀 처리

    labels = []
    for sentence in sentences:
        predicted_label = predict_sentiment(model, tokenizer, sentence)
        label_text = reverse_label_map[predicted_label]
        labels.append(label_text)

    # 'Label' 열 추가 또는 업데이트
    label_column = ws.max_column + 1
    ws.cell(row=1, column=label_column, value='Label')
    for i, label in enumerate(labels, start=2):
        ws.cell(row=i, column=label_column, value=label)

    # 변경 사항 저장
    wb.save(file_path)

# result_sub.xlsx에 감정 라벨 추가
result_sub_file_path = '../xlsx/result_sub.xlsx'  # 파일 경로에 맞게 수정해주세요
add_labels_to_excel(result_sub_file_path, tokenizer, model)

print("감정 라벨이 추가된 result_sub.xlsx 파일이 저장되었습니다.")
