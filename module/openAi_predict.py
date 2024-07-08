import openai
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, f1_score
from collections import Counter

# OpenAI API 키 설정
openai.api_key = 'myKey'
context = '사회적 물의를 일으킨 것의 대한 유튜버의 사과영상'
youtuber = '오킹'

# 감정 문자열을 OpenAI 엔진에 맞게 매핑
label_map = {'행복': 'happy', '혐오': 'disgust', '공포': 'fear', '슬픔': 'sadness', '놀람': 'surprise', '분노': 'anger', '중립': 'neutral'}
reverse_label_map = {v: k for k, v in label_map.items()}

def load_dataset(file_path):
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active  # 현재 활성화된 시트를 가져옴
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        exit(1)
    
    # 'sentence' 열만 가져옴
    data_list = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            sentence = str(row[0])
            data_list.append([sentence])
    
    return data_list[0:30]  # 처음 30개 데이터만 반환

# OpenAI API를 사용하여 감정 분석을 수행하는 함수 정의
def predict_sentiment(text):
    try:
        messages = [
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": f"'{context}' 문맥을 기반으로 '{text}'를 행복, 웃김, 슬픔, 분노, 중립 5가지 감정중 하나로 라벨링 후 감정 라벨을 반환하라. "}
        ]
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=messages,
            max_tokens=50,
            stop=["\n", "감정 분석 결과:"],
        )
        # OpenAI API의 응답에서 감정 라벨 추출
        predicted_label = response.choices[0].message['content'].strip().lower()
        return predicted_label
    except openai.error.APIError as e:
        print(f"OpenAI API 호출 오류: {e}")
        return "API 호출 오류"

# OpenAI API를 사용하여 감정 분석을 수행하는 함수 정의
def predict_object(text, num_predictions=5):
    try:
        messages = [
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": f"문맥: {context}. 유튜버: {youtuber}."},
            {"role": "user", "content": "다음 문장이 얘기하는 대상이 context, youtuber, other중 어떤 것인지 선택하세요. 예: '안드레진 미친놈이네' -> context"},
            {"role": "user", "content": f"문장: '{text}'"}
        ]
        results = []
        for _ in range(num_predictions):
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=messages,
                max_tokens=50,
            )
            predicted_label = response.choices[0].message['content'].strip().lower()
            results.append(predicted_label)
        most_common_label = Counter(results).most_common(1)[0][0]
        return most_common_label
    except openai.error.APIError as e:
        print(f"OpenAI API 호출 오류: {e}")
        return "API 호출 오류"

def evaluate_model(file_path):
    data_list = load_dataset(file_path)
    sentences = [item[0] for item in data_list]
    accuracy = 26/30
    predicted_emotions = []
    predicted_objects = []
    for sentence in sentences:
        predicted_emo = predict_sentiment(sentence)
        predicted_emotions.append(predicted_emo)
        predicted_obj = predict_object(sentence)
        predicted_objects.append(predicted_obj)
        print(f"원문: {sentence}, 예측 감정: {predicted_emo}, 대상: {predicted_obj}")  # 예측 결과 출력
    # Note: Evaluation metrics like accuracy and F1 score are removed here.

# 예측만 수행하고 싶은 경우:
valid_data_file_path = '../xlsx/crawling_auto.xlsx'  # 파일 경로에 맞게 수정해주세요
evaluate_model(valid_data_file_path)
