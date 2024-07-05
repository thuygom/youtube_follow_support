from transformers import BertForSequenceClassification, BertTokenizer
import torch
from openpyxl import load_workbook  # openpyxl 임포트

# 모델과 토크나이저 불러오기
model = BertForSequenceClassification.from_pretrained('./results')

# KoBERT 모델 및 토크나이저 로딩
model_name = "monologg/kobert"
tokenizer = BertTokenizer.from_pretrained(model_name)

print("체크포인트에서 모델과 토크나이저가 성공적으로 불러와졌습니다.")

# GPU가 있을 경우 활용
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)

# 모델 가중치 일부 확인
print("모델 가중치 확인:")
print(model.bert.encoder.layer[0].attention.self.query.weight[:5])  # 일부 가중치 출력

# Fine-tuning된 모델을 사용하여 감정 분석 수행 함수 정의
def predict_sentiment(model, tokenizer, text, max_length=128):
    model.eval()  # 모델을 평가 모드로 전환
    encoded_input = tokenizer(text, truncation=True, padding='max_length', max_length=max_length, return_tensors='pt')
    input_ids = encoded_input['input_ids'].to(device)
    attention_mask = encoded_input['attention_mask'].to(device)
    with torch.no_grad():
        output = model(input_ids=input_ids, attention_mask=attention_mask)
    predicted_label = torch.argmax(output.logits[0]).item()
    return predicted_label

# 단일 문장 테스트
test_sentence = "이 문장은 테스트 문장입니다."
predicted_label = predict_sentiment(model, tokenizer, test_sentence)
print(f"테스트 문장의 예측 라벨: {predicted_label}")

def add_labels_to_excel(file_path, tokenizer, model):
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        exit(1)

    # 데이터가 있는 열을 확인하고 데이터 읽기
    column_idx = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'Processed_Comments':  # 또는 다른 식별자를 사용할 수 있음
            column_idx = col
            break
    
    if column_idx is None:
        print("Cannot find 'Processed_Comments' column in the Excel file.")
        exit(1)
    
    sentences = []
    for row in ws.iter_rows(min_row=2, max_col=column_idx, values_only=True):
        if row[0] is not None:
            sentences.append(str(row[0]))
        else:
            sentences.append("")  # 빈 셀 처리

    labels = []
    for sentence in sentences:
        predicted_label = predict_sentiment(model, tokenizer, sentence)
        labels.append(predicted_label)
        print(f"문장: {sentence}, 예측 라벨: {predicted_label}")  # 예측 결과 출력

    # 'Label' 열 추가 또는 업데이트
    label_column = ws.max_column + 1
    ws.cell(row=1, column=label_column, value='Label')
    for i, label in enumerate(labels, start=2):
        ws.cell(row=i, column=label_column, value=label)

    # 변경 사항 저장
    wb.save(file_path)

# result_sub.xlsx에 영어 감정 라벨 추가하기
result_sub_file_path = '../xlsx/result_sub.xlsx'  # 파일 경로에 맞게 수정해주세요
add_labels_to_excel(result_sub_file_path, tokenizer, model)

print("감정 라벨이 추가된 result_sub.xlsx 파일이 저장되었습니다.")
