import torch
from transformers import BertForSequenceClassification, BertTokenizer
from sklearn.metrics import accuracy_score, f1_score
from openpyxl import load_workbook

# 모델과 토크나이저 불러오기
model = BertForSequenceClassification.from_pretrained('./results2')

# KoBERT 모델 및 토크나이저 로딩
model_name = "monologg/kobert"
tokenizer = BertTokenizer.from_pretrained(model_name)

print("체크포인트에서 모델과 토크나이저가 성공적으로 불러와졌습니다.")

# GPU가 있을 경우 활용
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)

# 모델 가중치 일부 확인
print("모델 가중치 확인:")
print(model.bert.encoder.layer[0].attention.self.query.weight[:5])  # 일부 가중치 출력

# 감정 문자열을 숫자로 매핑
label_map = {'행복': 0, '혐오': 1, '공포': 2, '슬픔': 3, '놀람': 4, '분노': 5, '중립': 6}
reverse_label_map = {v: k for k, v in label_map.items()}

def load_dataset(file_path):
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active  # 현재 활성화된 시트를 가져옴
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        exit(1)
    
    # 'sentence'와 'emotion' 열을 가져옴
    data_list = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            sentence = str(row[0])
            emotion = str(row[1])
            # 감정 문자열을 숫자로 매핑
            label = label_map[emotion]
            data_list.append([sentence, label])
    
    return data_list

# Fine-tuning된 모델을 사용하여 감정 분석 수행 함수 정의
def predict_sentiment(model, tokenizer, text, max_length=128):
    model.eval()  # 모델을 평가 모드로 전환
    encoded_input = tokenizer(text, truncation=True, padding='max_length', max_length=max_length, return_tensors='pt')
    input_ids = encoded_input['input_ids'].to(device)
    attention_mask = encoded_input['attention_mask'].to(device)
    with torch.no_grad():
        output = model(input_ids=input_ids, attention_mask=attention_mask)
    predicted_label = torch.argmax(output.logits[0]).item()
    return predicted_label

def evaluate_model(file_path, tokenizer, model):
    data_list = load_dataset(file_path)
    sentences = [item[0] for item in data_list]
    true_labels = [item[1] for item in data_list]

    predicted_labels = []
    for sentence, true_label in zip(sentences, true_labels):
        predicted_label = predict_sentiment(model, tokenizer, sentence)
        predicted_labels.append(predicted_label)
        print(f"문장: {sentence}, 예측 라벨: {reverse_label_map[predicted_label]}, 정답: {reverse_label_map[true_label]}")  # 예측 결과와 정답 출력

    # 성능 평가
    accuracy = accuracy_score(true_labels, predicted_labels)
    f1 = f1_score(true_labels, predicted_labels, average='weighted')

    print(f"Accuracy: {accuracy}")
    print(f"F1 Score: {f1}")

# valid_data.xlsx 파일로 평가
valid_data_file_path = 'valid_data.xlsx'  # 파일 경로에 맞게 수정해주세요
evaluate_model(valid_data_file_path, tokenizer, model)
