import pandas as pd
from openpyxl import load_workbook
import random

# 원본 엑셀 파일 경로
original_file_path1 = './train_data.xlsx'
original_file_path2 = './train_data2.xlsx'

# 샘플링할 데이터 수
num_samples_per_emotion = 10

# 새로운 엑셀 파일 경로
sampled_file_path = './valid_data.xlsx'

# 엑셀 파일에서 데이터셋 로딩 함수
def load_dataset(file_path):
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active  # 현재 활성화된 시트를 가져옴
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        exit(1)
    
    # 'sentence'와 'emotion' 열을 가져옴
    data_list = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            sentence = str(row[0])
            emotion = str(row[1])
            data_list.append((sentence, emotion))
    
    return data_list

# 감정별로 데이터를 섞고 샘플링하는 함수
def sample_data(data_list, num_samples_per_emotion):
    # 감정별로 데이터를 모으기
    emotion_data = {}
    for sentence, emotion in data_list:
        if emotion not in emotion_data:
            emotion_data[emotion] = []
        emotion_data[emotion].append((sentence, emotion))
    
    # 감정별로 데이터를 무작위로 섞은 후 샘플링
    sampled_data = []
    for emotion, data in emotion_data.items():
        random.shuffle(data)  # 데이터를 무작위로 섞음
        sampled_data.extend(data[:num_samples_per_emotion])  # 일정 개수만큼 샘플링
    
    # 데이터를 무작위로 섞음
    random.shuffle(sampled_data)
    
    return sampled_data

# 데이터셋 로딩
data_list1 = load_dataset(original_file_path1)
data_list2 = load_dataset(original_file_path2)

# 데이터 샘플링
sampled_data1 = sample_data(data_list1, num_samples_per_emotion=num_samples_per_emotion)
sampled_data2 = sample_data(data_list2, num_samples_per_emotion=num_samples_per_emotion)

# 데이터 병합
sampled_data_combined = sampled_data1 + sampled_data2

# 데이터프레임 생성
df = pd.DataFrame(sampled_data_combined, columns=['Sentence', 'Emotion'])

# 데이터프레임을 엑셀 파일로 저장
df.to_excel(sampled_file_path, index=False)

print(f"샘플링된 데이터가 {sampled_file_path}에 저장되었습니다.")
